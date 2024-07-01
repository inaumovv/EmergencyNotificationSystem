from contextlib import contextmanager
from io import BytesIO

import openpyxl
from openpyxl.cell import Cell


class FileParser:

    @contextmanager
    def __get_sheet(self, file: BytesIO):
        yield openpyxl.load_workbook(file).active

    def parse_file(self, file: BytesIO):
        contact_data: dict = {'phone_numbers': [], 'email_addresses': []}
        with self.__get_sheet(file) as sheet:
            for column in range(sheet.max_column):
                cell: Cell = sheet.cell(row=1, column=column + 1)
                if 'телефон' in cell.value.lower():
                    for row in range(2, sheet.max_row + 1):
                        phone_number: int = sheet.cell(row=row, column=column + 1).value
                        contact_data['phone_numbers'].append(phone_number)
                if 'почта' in cell.value.lower():
                    for row in range(2, sheet.max_row + 1):
                        email: str = sheet.cell(row=row, column=column + 1).value
                        contact_data['email_addresses'].append(email)
        return contact_data
